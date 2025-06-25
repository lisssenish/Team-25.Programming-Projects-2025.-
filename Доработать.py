import telebot
from telebot import types
import pandas as pd
from io import BytesIO
import os
import re
import time
import tempfile
from functools import wraps
from datetime import datetime
import threading
import pathlib as path
from openpyxl import load_workbook, Workbook

RIGHTS_FILE = '/content/drive/MyDrive/rights.xlsx'
REPORTS_DIR = "/content/drive/MyDrive/reports"
os.makedirs(REPORTS_DIR, exist_ok=True)
if os.path.exists('shopofmanagers.csv'):
    shopofmanagers = pd.read_csv('shopofmanagers.csv')
else:
    shopofmanagers = pd.DataFrame(columns=['id_manager', 'name_manager', 'id_shop', 'name_shop'])
def load_users():
    """Загрузка пользователей из файла"""
    if os.path.exists(RIGHTS_FILE):
        try:
            return pd.read_excel(RIGHTS_FILE)
        except:
            # Если файл поврежден, создаем новый
            columns = ['id', 'role', 'name', 'number']
            return pd.DataFrame(columns=columns)
    else:
        # Создаем файл с администратором по умолчанию
        default_admin = pd.DataFrame({
            'id': ['A00001'],
            'role': ['Administrator'],
            'name': ['IgorSchenderovich'],
            'number': [1]
        })
        save_users(default_admin)
        return default_admin

def save_users(df):
    """Сохранение пользователей в файл"""
    try:
        os.makedirs(os.path.dirname(RIGHTS_FILE), exist_ok=True)
        df.to_excel(RIGHTS_FILE, index=False, engine='openpyxl')
        return True
    except Exception as e:
        print(f"Ошибка сохранения пользователей: {e}")
        return False

names = load_users()
TOKEN = '7791429879:AAEgbCL8bFjQYnb81Rf1s2Hn_F5lRbZ3eKo'
bot = telebot.TeleBot(TOKEN)

# Словари для хранения состояний и ролей пользователей
user_states = {}
user_roles = {}  # Новый словарь для хранения ролей
shop_data = {}    # Словарь для временного хранения данных магазина
user_chat_id={}
shop = {}
#Стартовые функции
@bot.message_handler(commands=['start', 'help'])
def welcome_message(message):
    welcome_text = """
Добро пожаловать в бота для автоматизации отчётов!
Этот бот умеет:
- Собирать информацию с точек
- Предоставлять отчёты
- Систематизировать информацию

Введите команду /id для начала работы
    """
    bot.send_message(message.chat.id, welcome_text)






@bot.message_handler(commands=['id'])
def handle_id_command(message):
    msg = bot.send_message(message.chat.id,
                           "🔑 Введите ваш ID в формате: Префикс (A/M/S) + 5 цифр\nПример: A00123 или M00001")
    bot.register_next_step_handler(msg, process_user_id)
def process_user_id(message):
    try:
        user_id = message.text.strip().upper()
        role_mapping = {'A': 'admin', 'M': 'manager', 'S': 'shop'}

        # Валидация формата ID
        if len(user_id) != 6 or user_id[0] not in role_mapping or not user_id[1:].isdigit():
            bot.reply_to(message, "❌ Некорректный формат ID. Пример: A00123")
            return

        # Поиск пользователя в базе
        user_row = names.loc[names['id'] == user_id]

        if user_row.empty:
            bot.reply_to(message, "⛔ Пользователь с таким ID не зарегистрирован в системе")
            return

        # Определение и сохранение роли
        role_prefix = user_id[0]
        role = role_mapping.get(role_prefix)

        user_roles[message.chat.id] = role
        user_chat_id[message.chat.id]=user_id

        # Формирование ответа
        responses = {
            'admin': "👑 Администратор\nДоступные команды:\n/stats - статистика\n/users - управление пользователями",
            'manager': "📊 Менеджер\nДоступные команды:\n/plan - установить план\n/statm - статистика\n/my_shops-ваши точки",
            'shop': "🏪 Магазин\nДоступные команды:\n/report - сдать отчет"
        }

        bot.send_message(message.chat.id, f"✅ Авторизация успешна!\n{responses[role]}")

    except KeyError:
        bot.reply_to(message, "⚠️ Неизвестный тип аккаунта. Обратитесь к администратору")
    except Exception as e:
        bot.reply_to(message, f"🚨 Критическая ошибка: {str(e)}")





# Функции для администратора


#Декораторы
def admin_required(func):
    @wraps(func)
    def wrapper(message, *args, **kwargs):
        if user_roles.get(message.chat.id) != 'admin':
            bot.reply_to(message, "⛔ Требуется роль администратора!")
            return
        return func(message, *args, **kwargs)

    return wrapper





def error_handler(func):
    @wraps(func)
    def wrapper(message, *args, **kwargs):
        try:
            return func(message, *args, **kwargs)
        except Exception as e:
            handle_error(message, e)

    return wrapper


def validate_id(id_str: str, role: str = None) -> str:
    """Валидация ID с обязательной заглавной буквой в префиксе"""
    if len(id_str) != 6:
        raise ValueError(f"Некорректная длина ID: {len(id_str)} (требуется 6 символов)")

    prefix = id_str[0]
    number_part = id_str[1:]

    # Проверка префикса
    valid_prefixes = {'A', 'M', 'S'}
    if prefix not in valid_prefixes:
        raise ValueError(f"Неверный префикс. Допустимые: {', '.join(valid_prefixes)}")

    # Проверка роли
    if role:
        expected_prefix = role[0].upper()
        if prefix != expected_prefix:
            raise ValueError(f"Для роли {role} требуется префикс {expected_prefix}")

    # Проверка числовой части
    if not number_part.isdigit():
        raise ValueError("После префикса должны быть 5 цифр")

    return id_str  # Возвращаем оригинальный ID


def find_user(id_str: str, role: str) -> pd.Series:
    """Поиск пользователя с учетом регистра"""
    # Исправлено название столбца 'id' вместо 'user id'
    df = names[
        (names['id'] == id_str) &
        (names['role'].str.upper() == role.upper())
        ]

    if df.empty:
        raise ValueError(f"{role.capitalize()} с ID {id_str} не найден")

    return df.iloc[0]





#Функция link_shop_to_manager
@bot.message_handler(commands=['link_shop_to_manager'])
@admin_required
@error_handler
def start_linking(message):
    user_data[message.chat.id] = {'step': 'manager'}
    bot.send_message(message.chat.id, "🔗 Введите ID менеджера (формат: M12345):")


@bot.message_handler(func=lambda m: user_data.get(m.chat.id, {}).get('step') == 'manager')
@admin_required
@error_handler
def process_manager(message):
    chat_id = message.chat.id
    try:
        raw_input = message.text.strip().upper()
        manager_id = validate_id(raw_input, 'manager')
        manager = find_user(manager_id, 'Manager')

        user_data[chat_id] = {
            'manager': {
                'id': manager_id,
                'name': manager['name']
            },
            'shops': [],
            'step': 'shop'
        }
        bot.send_message(chat_id, "🏪 Введите ID магазина (формат: S12345):")

    except ValueError as ve:
        bot.send_message(chat_id, f"❌ {str(ve)}")
        start_linking(message)


@bot.message_handler(func=lambda m: user_data.get(m.chat.id, {}).get('step') == 'shop')
@admin_required
@error_handler
def process_shop(message):
    chat_id = message.chat.id
    try:
        raw_input = message.text.strip().upper()
        shop_id = validate_id(raw_input, 'shop')
        shop = find_user(shop_id, 'Shop')

        # Проверка существующей связи
        exists = not shopofmanagers[
            (shopofmanagers['id_manager'] == user_data[chat_id]['manager']['id']) &
            (shopofmanagers['id_shop'] == shop_id)
            ].empty

        if exists:
            raise ValueError("Связь уже существует")

        user_data[chat_id]['shops'].append({
            'id': shop_id,
            'name': shop['name']
        })

        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.add("➕ Добавить ещё магазин", "✅ Завершить привязку")

        bot.send_message(
            chat_id,
            f"✅ Успешно привязано к:\n{shop['name']}",
            reply_markup=markup
        )
        user_data[chat_id]['step'] = 'confirmation'

    except ValueError as ve:
        bot.send_message(chat_id, f"❌ {str(ve)}\nПовторите ввод ID магазина:")


@bot.message_handler(func=lambda m: user_data.get(m.chat.id, {}).get('step') == 'confirmation')
@admin_required
@error_handler
def finalize_linking(message):
    chat_id = message.chat.id
    data = user_data[chat_id]

    if message.text == '➕ Добавить ещё магазин':
        data['step'] = 'shop'
        bot.send_message(chat_id, "🏪 Введите ID следующего магазина:")
        return

    # Сохранение всех связей
    new_entries = [{
        'id_manager': data['manager']['id'],
        'name_manager': data['manager']['name'],
        'id_shop': shop['id'],
        'name_shop': shop['name']
    } for shop in data['shops']]
    try:
        pd.DataFrame(new_entries).to_csv(
            'shopofmanagers.csv',
            mode='a',
            header=not pd.io.common.file_exists('shopofmanagers.csv'),
            index=False
        )

        report = [
                     f"🔗 Создано связей: {len(new_entries)}",
                     f"👨💼 Менеджер: {data['manager']['name']}",
                     "🏪 Магазины:"
                 ] + [f"• {shop['name']}" for shop in data['shops']]

        bot.send_message(
            chat_id,
            "\n".join(report),
            reply_markup=types.ReplyKeyboardRemove()
        )

    finally:
        if chat_id in user_data:
            del user_data[chat_id]


def handle_error(message, error):
    chat_id = message.chat.id
    error_msg = [
        "⚠️ Произошла ошибка:",
        f"• Тип: {type(error).__name__}",
        f"• Описание: {str(error)}"
    ]

    bot.send_message(chat_id, "\n".join(error_msg))
    bot.send_message(chat_id, "🔄 Сессия сброшена. Начните заново.")

    if chat_id in user_data:
        del user_data[chat_id]






#Функция get_links
@bot.message_handler(commands=['get_links'])
@admin_required
@error_handler
def get_links_command(message):
    """
    Отправка файла связей в формате Excel
    """
    chat_id = message.chat.id

    try:
        # Всегда читаем актуальный файл
        if not os.path.exists('shopofmanagers.csv'):
            bot.send_message(chat_id, "📭 База связей пуста")
            return

        df = pd.read_csv('shopofmanagers.csv')

        expected_columns = ['id_manager', 'name_manager', 'id_shop', 'name_shop']
        if not all(col in df.columns for col in expected_columns):
            missing = set(expected_columns) - set(df.columns)
            raise ValueError(f"Отсутствуют колонки: {', '.join(missing)}")

        report_data = df[expected_columns]

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            with pd.ExcelWriter(tmp.name, engine='xlsxwriter') as writer:
                report_data.to_excel(
                    writer,
                    sheet_name='Связи',
                    index=False,
                    header=['ID менеджера', 'Имя менеджера', 'ID магазина', 'Название магазина']
                )

                worksheet = writer.sheets['Связи']
                for col_num, col_name in enumerate(expected_columns):
                    max_len = max(
                        report_data[col_name].astype(str).map(len).max(),
                        len(col_name)
                    ) + 2
                    worksheet.set_column(col_num, col_num, max_len)

            with open(tmp.name, 'rb') as file:
                bot.send_document(
                    chat_id=chat_id,
                    document=file,
                    caption='🔗 Актуальные связи менеджеров и магазинов',
                    visible_file_name='Manager_Shop_Links.xlsx'
                )

    except Exception as e:
        bot.send_message(chat_id, f"⚠️ Ошибка генерации отчета: {str(e)}")
    finally:
        if 'tmp' in locals():
            try:
                os.remove(tmp.name)
            except:
                pass


#Функция stats
@bot.message_handler(commands=['stats'])
def handle_stats(message):
    if user_roles.get(message.chat.id) == 'admin':
        stats_text = """
📊 Статистика системы:
- Всего пользователей: 143
- Активных магазинов: 67
- Выполнение плана: 82%
        """
        bot.send_message(message.chat.id, stats_text)
    else:
        bot.send_message(message.chat.id, "⛔ Доступ запрещён! Требуются права администратора")

@bot.message_handler(commands=['reports'])
def handle_reports_command(message):
    try:
        # Проверка прав доступа
        if user_roles.get(message.chat.id) != 'admin':
            bot.send_message(message.chat.id, "⛔ Требуются права администратора!")
            return

        # Создаем клавиатуру с годами
        markup = types.ReplyKeyboardMarkup(one_time_keyboard=True, resize_keyboard=True)
        current_year = datetime.now().year
        markup.add(str(current_year), str(current_year - 1))
        
        # Сохраняем состояние
        user_data[message.chat.id] = {'step': 'select_year'}
        
        bot.send_message(message.chat.id, "📅 Выберите год:", reply_markup=markup)
        bot.register_next_step_handler(message, process_year_selection)

    except Exception as e:
        bot.send_message(message.chat.id, f"❌ Ошибка: {str(e)}")

def process_year_selection(message):
    try:
        chat_id = message.chat.id
        
        if message.text not in [str(datetime.now().year), str(datetime.now().year - 1)]:
            raise ValueError("Недопустимый год")
            
        # Сохраняем год и переходим к выбору месяца
        user_data[chat_id] = {
            'step': 'select_month',
            'year': int(message.text)
        }
        
        # Создаем клавиатуру с месяцами
        markup = types.ReplyKeyboardMarkup(one_time_keyboard=True, resize_keyboard=True, row_width=3)
        months = [
            "1️⃣ Январь", "2️⃣ Февраль", "3️⃣ Март",
            "4️⃣ Апрель", "5️⃣ Май", "6️⃣ Июнь",
            "7️⃣ Июль", "8️⃣ Август", "9️⃣ Сентябрь",
            "🔟 Октябрь", "1️⃣1️⃣ Ноябрь", "1️⃣2️⃣ Декабрь"
        ]
        markup.add(*months)
        
        bot.send_message(chat_id, "📅 Выберите месяц:", reply_markup=markup)
        bot.register_next_step_handler(message, process_month_selection)

    except Exception as e:
        bot.send_message(message.chat.id, f"❌ Ошибка: {str(e)}")
        if message.chat.id in user_data:
            del user_data[message.chat.id]

def process_month_selection(message):
    try:
        chat_id = message.chat.id
        user_info = user_data.get(chat_id, {})
        
        if user_info.get('step') != 'select_month' or 'year' not in user_info:
            raise ValueError("Неверная последовательность команд")
            
        # Получаем номер месяца из emoji
        month_map = {
            "1️⃣": 1, "2️⃣": 2, "3️⃣": 3, "4️⃣": 4, "5️⃣": 5, "6️⃣": 6,
            "7️⃣": 7, "8️⃣": 8, "9️⃣": 9, "🔟": 10, "1️⃣1️⃣": 11, "1️⃣2️⃣": 12
        }
        
        month_emoji = message.text.split()[0]
        month_num = month_map.get(month_emoji)
        
        if not month_num:
            raise ValueError("Недопустимый месяц")
            
        # Формируем имя файла
        file_name = f"reports_{user_info['year']}_{month_num:02d}.xlsx"
        file_path = os.path.join(REPORTS_DIR, file_name)

        if not os.path.exists(file_path):
            bot.send_message(chat_id, f"❌ Файл отчетов за указанный период не найден!", 
                          reply_markup=types.ReplyKeyboardRemove())
            return

        # Получаем название месяца
        month_names = [
            "январь", "февраль", "март", "апрель", "май", "июнь",
            "июль", "август", "сентябрь", "октябрь", "ноябрь", "декабрь"
        ]
        month_name = month_names[month_num - 1]
        
        # Читаем и анализируем файл
        df = pd.read_excel(file_path, sheet_name=None)
        sheets_count = len(df)
        shops_count = sum(len(sheet) for sheet in df.values() if isinstance(sheet, pd.DataFrame))
        
        # Форматируем сообщение
        caption = (
            f"📊 Отчет за {month_name} {user_info['year']} года\n"
            f"📂 Файл содержит {sheets_count} листов\n"
            f"🏪 Всего отчетов: {shops_count}"
        )

        # Отправляем файл с подписью
        with open(file_path, 'rb') as file:
            bot.send_document(
                chat_id=chat_id,
                document=file,
                caption=caption,
                reply_markup=types.ReplyKeyboardRemove(),
                visible_file_name=f"Отчеты_{month_name}_{user_info['year']}.xlsx"
            )
            
    except Exception as e:
        bot.send_message(chat_id, f"❌ Ошибка: {str(e)}")
    finally:
        if chat_id in user_data:
            del user_data[chat_id]

@bot.message_handler(commands=['reports_current'])
def handle_reports_current(message):
    """Отправка отчета за текущий месяц"""
    try:
        if user_roles.get(message.chat.id) != 'admin':
            bot.send_message(message.chat.id, "⛔ Требуются права администратора!")
            return

        current_date = datetime.now()
        file_name = f"reports_{current_date.year}_{current_date.month:02d}.xlsx"
        file_path = os.path.join(REPORTS_DIR, file_name)

        if not os.path.exists(file_path):
            bot.send_message(message.chat.id, "❌ Файл отчетов за текущий месяц не найден!")
            return

        # Читаем файл для анализа
        df = pd.read_excel(file_path, sheet_name=None)
        sheets_count = len(df)
        shops_count = sum(len(sheet) for sheet in df.values() if isinstance(sheet, pd.DataFrame))
        
        month_names = [
            "январь", "февраль", "март", "апрель", "май", "июнь",
            "июль", "август", "сентябрь", "октябрь", "ноябрь", "декабрь"
        ]
        month_name = month_names[current_date.month - 1]

        caption = (
            f"📊 Отчет за текущий месяц ({month_name} {current_date.year})\n"
            f"📂 Файл содержит {sheets_count} листов\n"
            f"🏪 Всего отчетов: {shops_count}"
        )

        with open(file_path, 'rb') as file:
            bot.send_document(
                chat_id=message.chat.id,
                document=file,
                caption=caption,
                visible_file_name=f"Отчеты_текущий_{current_date.strftime('%Y-%m')}.xlsx"
            )

    except Exception as e:
        bot.send_message(message.chat.id, f"❌ Ошибка: {str(e)}")





#Функция users
@bot.message_handler(commands=['users'])
def handle_users(message):
    if user_roles.get(message.chat.id) == 'admin':
        users_text = """
👥 Управление пользователями:
Команды:
/get_info - вывод всех зарегестрированных аккаунтов
/get_names- посмотреть список всех пользователей(xls)
/add_user - добавить пользователя
/remove_user - удалить пользователя
/link_shop_to_manager - привязка магазина к менеджеру
/get_links - получение файла со связями
        """
        bot.send_message(message.chat.id, users_text)
    else:
        bot.send_message(message.chat.id, "⛔ Доступ запрещён! Требуются права администратора")

user_data = {}
def create_and_add_id(role, name):
    roles_index = {'Developer': 'T', 'Administrator': 'A', 'Manager': 'M', 'Shop': 'S'}  # Заглавные буквы
    k = roles_index[role]

    existing = names[names['role'] == role]
    if existing.empty:
        n = 1
    else:
        n = existing['number'].max() + 1

    numeric_part = str(n).zfill(5)
    new_id = f"{k}{numeric_part}"  # Теперь ID с заглавной буквы
    names.loc[len(names)] = [new_id, role, name, n]
    names.to_csv('rights.xlsx', index=False)
    return new_id
def clean_user_data(chat_id):
    """Удаляем временные данные пользователя"""
    if chat_id in user_data:
        del user_data[chat_id]





#Функция add_user
@bot.message_handler(commands=['add_user'])
def handle_add_user(message):
    if user_roles.get(message.chat.id) != 'admin':
        bot.send_message(message.chat.id, "⛔ Требуются права администратора!")
        return

    markup = types.ReplyKeyboardMarkup(one_time_keyboard=True, resize_keyboard=True, row_width=2)
    markup.add('Shop', 'Administrator', 'Manager')

    bot.send_message(message.chat.id, "🔽 Выберите роль из кнопок ниже:", reply_markup=markup)
    bot.register_next_step_handler(message, process_role_step)

def process_role_step(message):
    try:
        chat_id = message.chat.id
        role_mapping = {
            'shop': 'Shop',
            'administrator': 'Administrator',
            'manager': 'Manager'
        }
        
        input_role = message.text.strip().lower()
        role_name = role_mapping.get(input_role)
        
        if not role_name:
            raise ValueError("Некорректная роль. Выберите из предложенных вариантов.")

        user_data[chat_id] = {'role': role_name}
        bot.send_message(chat_id, "✏️ Введите название точки:", reply_markup=types.ReplyKeyboardRemove())
        bot.register_next_step_handler(message, process_location_step)

    except Exception as e:
        bot.reply_to(message, f"❌ Ошибка: {str(e)}")
        clean_user_data(chat_id)

def process_location_step(message):
    try:
        chat_id = message.chat.id
        if message.content_type != 'text':
            raise ValueError("Название точки должно быть текстом")

        location = message.text.strip()
        if len(location) < 2:
            raise ValueError("Слишком короткое название точки (минимум 2 символа)")

        role_data = user_data.get(chat_id, {})
        role_name = role_data.get('role')

        if not role_name:
            raise ValueError("Сессия устарела. Начните заново.")

        new_user_id = create_user_id(role_name)
        
        global names
        new_user = pd.DataFrame({
            'id': [new_user_id],
            'role': [role_name],
            'name': [location],
            'number': [get_next_number(role_name)]
        })
        
        names = pd.concat([names, new_user], ignore_index=True)
        
        if save_users(names):
            bot.send_message(chat_id, f"✅ Пользователь создан!\nID: {new_user_id}\nТочка: {location}")
        else:
            raise Exception("Не удалось сохранить пользователя")

    except Exception as e:
        bot.reply_to(message, f"❌ Ошибка: {str(e)}")
    finally:
        clean_user_data(chat_id)

def create_user_id(role):
    prefix_map = {
        'Administrator': 'A',
        'Manager': 'M',
        'Shop': 'S'
    }
    prefix = prefix_map.get(role, 'U')
    max_num = names[names['role'] == role]['number'].max()
    next_num = 1 if pd.isna(max_num) else max_num + 1
    return f"{prefix}{next_num:05d}"

def get_next_number(role):
    max_num = names[names['role'] == role]['number'].max()
    return 1 if pd.isna(max_num) else max_num + 1

def clean_user_data(chat_id):
    if chat_id in user_data:
        del user_data[chat_id]







#Функция get_names
@bot.message_handler(commands=['get_names'])
def send_names_excel(message):
    try:
        if user_roles.get(message.chat.id) != 'admin':
            bot.reply_to(message, "⛔ Требуются права администратора!")
            return

        if names.empty:
            bot.reply_to(message, "📭 База данных пуста")
            return

        output = BytesIO()

        # Используем правильный параметр sheet_name
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            names.to_excel(
                excel_writer=writer,
                sheet_name='Users',  # Правильное написание
                index=False
            )

        output.seek(0)
        bot.send_document(
            chat_id=message.chat.id,
            document=output,
            caption='📊 Полная база данных пользователей',
            visible_file_name='users_database.xlsx'
        )

    except Exception as e:
        bot.reply_to(message, f"❌ Ошибка при генерации файла: {str(e)}")
    finally:
        output.close() if 'output' in locals() else None







#Функция remove_user
@bot.message_handler(commands=['remove_user'])
def handle_remove_user(message):
    if user_roles.get(message.chat.id) == 'admin':
        bot.send_message(
            message.chat.id,
            "✏️ Введите ID пользователя для удаления:",
            reply_markup=types.ReplyKeyboardRemove()
        )
        bot.register_next_step_handler(message, process_remove_user)
    else:
        bot.send_message(message.chat.id, "⛔ Требуются права администратора!")


def process_remove_user(message):
    try:
        chat_id = message.chat.id
        raw_input = message.text.strip()
        user_id_to_remove = raw_input[0].upper() + raw_input[1:].lower()

        if not re.match(r'^[A-Za-z]\d{5}$', user_id_to_remove):
            raise ValueError("Неверный формат ID. Пример: A12345")

        # Поиск без учета регистра
        mask = names['id'].str.upper() == user_id_to_remove.upper()

        if not mask.any():
            raise ValueError("Пользователь с таким ID не найден")

        # Получаем ID перед удалением
        found_id = names.loc[mask, 'id'].values[0]

        # Удаляем запись
        names.drop(names[mask].index, inplace=True)
        names.to_csv('rights.xlsx', index=False)

        # Используем сохраненный ID
        bot.send_message(chat_id, f"✅ Пользователь {found_id} удалён")

    except Exception as e:
        bot.reply_to(message, f"❌ Ошибка: {str(e)}")


def escape_markdown(text):
    escape_chars = '_*[]()~`>#+-=|{}.!'
    return ''.join(['\\' + char if char in escape_chars else char for char in text])


def safe_escape_markdown(text: str) -> str:
    """Экранирование символов MarkdownV2 с проверкой типа"""
    if not isinstance(text, str):
        raise TypeError("Input must be string")
    return escape_markdown(text)







#Функция get_info
@bot.message_handler(commands=['get_info'])
def handle_get_info(message: types.Message):
    try:
        # Проверка прав администратора
        if user_roles.get(message.chat.id) != 'admin':
            bot.reply_to(message, "⛔ Требуются права администратора!")
            return

        # Проверка на пустую базу
        if names.empty:
            bot.reply_to(message, "📭 База данных пуста")
            return

        # Конфигурация ролей
        ROLE_CONFIG = {
            'Administrator': {'emoji': '👑', 'name': 'Администраторы'},
            'Manager': {'emoji': '💼', 'name': 'Менеджеры'},
            'Shop': {'emoji': '🏪', 'name': 'Магазины'},
        }

        grouped = names.groupby('role')

        for role, config in ROLE_CONFIG.items():
            if role not in grouped.groups:
                continue

            role_users = grouped.get_group(role)

            # Заголовок роли
            role_header = (
                f"{config['emoji']} "
                f"*{safe_escape_markdown(config['name'])}* "
                f"[{len(role_users)}]"
            )
            bot.send_message(
                message.chat.id,
                role_header,
                parse_mode='MarkdownV2'
            )

            # Данные пользователей
            for _, user in role_users.iterrows():
                user_info = (
                    f"🆔 ID: `{safe_escape_markdown(str(user['id']))}`\n"
                    f"📌 Название: {safe_escape_markdown(str(user['name']))}\n"
                )
                bot.send_message(
                    message.chat.id,
                    user_info,
                    parse_mode='MarkdownV2'
                )
                time.sleep(0.2)

    except Exception as e:
        bot.reply_to(message, f"❌ Ошибка: {str(e)}")






#Функция remove_user
@bot.message_handler(commands=['remove_link'])
@admin_required
@error_handler
def start_remove_link(message):
    chat_id = message.chat.id
    user_data[chat_id] = {'step': 'remove_manager_id'}
    bot.send_message(chat_id, "🔍 Введите ID менеджера для удаления связей:")


@bot.message_handler(func=lambda m: user_data.get(m.chat.id, {}).get('step') == 'remove_manager_id')
@admin_required
@error_handler
def process_remove_manager(message):
    chat_id = message.chat.id
    user_data[chat_id]['manager_id'] = message.text.strip()
    user_data[chat_id]['step'] = 'remove_shop_id'
    bot.send_message(chat_id, "🔍 Теперь введите ID магазина для удаления связи:")


@bot.message_handler(func=lambda m: user_data.get(m.chat.id, {}).get('step') == 'remove_shop_id')
@admin_required
@error_handler
def process_remove_shop(message):
    chat_id = message.chat.id
    data = user_data[chat_id]

    try:
        # Чтение и фильтрация данных
        df = pd.read_csv('shopofmanagers.csv')
        initial_count = len(df)

        # Фильтрация записей
        mask = (df['id_manager'] == data['manager_id']) & (df['id_shop'] == message.text.strip())
        df = df[~mask]

        if len(df) == initial_count:
            raise ValueError("Связь не найдена")

        # Сохранение изменений
        df.to_csv('shopofmanagers.csv', index=False)

        # Формирование отчета
        report = [
            "✅ Связь успешно удалена:",
            f"👨💼 ID менеджера: {data['manager_id']}",
            f"🏪 ID магазина: {message.text.strip()}"
        ]

        bot.send_message(chat_id, "\n".join(report))

    except Exception as e:
        handle_error(message, e)
    finally:
        if chat_id in user_data:
            del user_data[chat_id]

#Функции менеджера
@bot.message_handler(commands=['my_shops'])
def handle_my_shops(message):
    user_tg_id = message.chat.id  # Telegram ID пользователя (число)

    # Проверка роли
    if user_roles.get(user_tg_id) != 'manager':
        bot.send_message(user_tg_id, "⛔ Доступ запрещён! Требуются права менеджера")
        return

    # Получаем user_id менеджера (например, "M00001") из словаря user_chat_id
    try:
        manager_id = user_chat_id[user_tg_id]  # Здесь получаем "M00001"
    except KeyError:
        bot.send_message(user_tg_id, "❌ Ваш аккаунт менеджера не привязан к системе")
        return

    try:
        # Фильтруем магазины по manager_id ("M00001")
        if isinstance(shopofmanagers, dict):
            manager_shops = [shop for shop in shopofmanagers.values() if str(shop.get('id_manager')) == manager_id]
        elif hasattr(shopofmanagers, 'iterrows'):  # pandas DataFrame
            manager_shops = [row.to_dict() for _, row in shopofmanagers.iterrows() if
                             str(row['id_manager']) == manager_id]
        else:  # Список или другой тип
            manager_shops = [shop for shop in shopofmanagers if str(shop.get('id_manager')) == manager_id]

        # Формируем ответ
        if not manager_shops:
            bot.send_message(user_tg_id, "ℹ️ К вам не привязано ни одного магазина")
            return

        response = "🏪 Ваши магазины:\n\n"
        for i, shop in enumerate(manager_shops, 1):
            response += (
                f"{i}. ID магазина: {shop.get('id_shop', 'N/A')}\n"
                f"   Название: {shop.get('name_shop', 'N/A')}\n\n"
            )

        bot.send_message(user_tg_id, response)

    except Exception as e:
        bot.send_message(user_tg_id, f"❌ Ошибка при получении списка магазинов: {str(e)}")
@bot.message_handler(commands=['set_plan'])
def handle_set_plan(message):
    if user_roles.get(message.chat.id) != 'manager':
        bot.send_message(message.chat.id, "⛔ Доступ запрещён! Требуются права менеджера")
        return

    try:
        manager_id = user_chat_id[message.chat.id]
        # Получаем список магазинов менеджера
        manager_shops = shopofmanagers[shopofmanagers['id_manager'] == manager_id]

        if manager_shops.empty:
            bot.send_message(message.chat.id, "ℹ️ К вам не привязано ни одного магазина")
            return

        # Создаем клавиатуру с магазинами
        markup = types.ReplyKeyboardMarkup(one_time_keyboard=True, resize_keyboard=True)
        for _, shop in manager_shops.iterrows():
            markup.add(f"{shop['id_shop']} - {shop['name_shop']}")

        msg = bot.send_message(
            message.chat.id,
            "🏪 Выберите магазин для установки плана:",
            reply_markup=markup
        )
        bot.register_next_step_handler(msg, process_shop_selection_for_plan)

    except Exception as e:
        bot.send_message(message.chat.id, f"❌ Ошибка: {str(e)}")

def process_shop_selection_for_plan(message):
    try:
        shop_info = message.text.split(' - ')[0]
        shop_id = shop_info.strip()

        # Проверяем что магазин существует и принадлежит менеджеру
        manager_id = user_chat_id[message.chat.id]
        valid_shops = shopofmanagers[
            (shopofmanagers['id_manager'] == manager_id) &
            (shopofmanagers['id_shop'] == shop_id)
        ]

        if valid_shops.empty:
            bot.send_message(message.chat.id, "❌ Магазин не найден или не принадлежит вам")
            return

        shop_name = valid_shops.iloc[0]['name_shop']

        # Сохраняем данные во временном хранилище
        user_data[message.chat.id] = {
            'action': 'set_plan',
            'shop_id': shop_id,
            'shop_name': shop_name
        }

        bot.send_message(
            message.chat.id,
            f"✏️ Введите сумму плана для магазина {shop_name}:",
            reply_markup=types.ReplyKeyboardRemove()
        )
        bot.register_next_step_handler(message, process_plan_amount)

    except Exception as e:
        bot.send_message(message.chat.id, f"❌ Ошибка: {str(e)}")

def process_plan_amount(message):
    try:
        amount = float(message.text.strip())
        chat_id = message.chat.id
        user_info = user_data.get(chat_id, {})

        if user_info.get('action') != 'set_plan':
            raise ValueError("Неверный контекст операции")

        shop_id = user_info['shop_id']
        shop_name = user_info['shop_name']
        manager_id = user_chat_id[chat_id]

        # Добавляем или обновляем план
        global plans_df
        if shop_id in plans_df['shop_id'].values:
            plans_df.loc[plans_df['shop_id'] == shop_id, ['plan_amount', 'date']] = [amount, datetime.datetime.now().date()]
        else:
            new_plan = pd.DataFrame([{
                'shop_id': shop_id,
                'shop_name': shop_name,
                'plan_amount': amount,
                'manager_id': manager_id,
                'date': datetime.datetime.now().date()
            }])
            plans_df = pd.concat([plans_df, new_plan], ignore_index=True)

        # Сохраняем в файл
        plans_df.to_csv('plans.csv', index=False)

        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.add("✅ Да", "❌ Нет")

        bot.send_message(
            chat_id,
            f"Подтвердите установку плана:\n"
            f"🏪 Магазин: {shop_name}\n"
            f"💰 Сумма плана: {amount} руб.\n\n"
            f"Всё верно?",
            reply_markup=markup
        )
        bot.register_next_step_handler(message, confirm_plan_setting)

    except ValueError:
        bot.send_message(message.chat.id, "❌ Введите корректную сумму (число)")
    except Exception as e:
        bot.send_message(message.chat.id, f"❌ Ошибка: {str(e)}")

def confirm_plan_setting(message):
    try:
        chat_id = message.chat.id
        if message.text == '✅ Да':
            bot.send_message(
                chat_id,
                "✅ План успешно установлен!",
                reply_markup=types.ReplyKeyboardRemove()
            )
        else:
            bot.send_message(
                chat_id,
                "❌ Установка плана отменена",
                reply_markup=types.ReplyKeyboardRemove()
            )

    except Exception as e:
        bot.send_message(message.chat.id, f"❌ Ошибка: {str(e)}")
    finally:
        if chat_id in user_data:
            del user_data[chat_id]

@bot.message_handler(commands=['get_plans'])
def handle_get_plans(message):
    if user_roles.get(message.chat.id) != 'manager':
        bot.send_message(message.chat.id, "⛔ Доступ запрещён! Требуются права менеджера")
        return

    try:
        manager_id = user_chat_id[message.chat.id]

        # Получаем планы только для магазинов этого менеджера
        manager_shops = shopofmanagers[shopofmanagers['id_manager'] == manager_id]['id_shop']
        manager_plans = plans_df[plans_df['shop_id'].isin(manager_shops)]

        if manager_plans.empty:
            bot.send_message(message.chat.id, "ℹ️ У вас нет установленных планов для магазинов")
            return

        # Создаем Excel файл
        output = BytesIO()
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            manager_plans.to_excel(writer, sheet_name='Plans', index=False)

            # Форматирование
            worksheet = writer.sheets['Plans']
            for i, col in enumerate(manager_plans.columns):
                width = max(manager_plans[col].astype(str).map(len).max(), len(col)) + 2
                worksheet.set_column(i, i, width)

        output.seek(0)

        bot.send_document(
            message.chat.id,
            output,
            caption='📊 Планы для ваших магазинов',
            visible_file_name=f'plans_{datetime.datetime.now().date()}.xlsx'
        )
        output.close()

    except Exception as e:
        bot.send_message(message.chat.id, f"❌ Ошибка при создании отчета: {str(e)}")

@bot.message_handler(commands=['statm'])
def handle_users(message):
    if user_roles.get(message.chat.id) == 'manager':
        users_text = """
План магазина А: 87% 1009290202 из 10019189101
План магазина Б: 87% 10002 из 10019189101
        """
        bot.send_message(message.chat.id, users_text)
    else:
        bot.send_message(message.chat.id, "⛔ Доступ запрещён! Требуются права менеджера")

# Функции для магазина
@bot.message_handler(commands=['my_plan'])
def handle_my_plan(message):
    if user_roles.get(message.chat.id) != 'shop':
        bot.send_message(message.chat.id, "⛔ Доступ запрещён! Требуются права магазина")
        return

    try:
        shop_id = user_chat_id[message.chat.id]
        shop_plan = plans_df[plans_df['shop_id'] == shop_id]

        if shop_plan.empty:
            bot.send_message(message.chat.id, "ℹ️ Для вашего магазина не установлен план")
            return

        plan = shop_plan.iloc[0]
        bot.send_message(
            message.chat.id,
            f"📊 План для вашего магазина:\n\n"
            f"🏪 Магазин: {plan['shop_name']}\n"
            f"💰 Сумма плана: {plan['plan_amount']} руб.\n"
            f"📅 Дата установки: {plan['date']}"
        )

    except Exception as e:
        bot.send_message(message.chat.id, f"❌ Ошибка: {str(e)}")
@bot.message_handler(commands=['report'])
def handle_report(message):
    if user_roles.get(message.chat.id) != 'shop':
        bot.send_message(message.chat.id, "⛔ Эта команда доступна только для магазинов")
        return

    try:
        # Получаем название магазина из базы данных по ID пользователя
        shop_id = user_chat_id.get(message.chat.id)
        if not shop_id:
            raise ValueError("Ваш магазин не зарегистрирован в системе")

        shop_info = names[names['id'] == shop_id]
        if shop_info.empty:
            raise ValueError("Информация о вашем магазине не найдена")

        shop_name = shop_info.iloc[0]['name']
        
        # Инициализируем данные для отчета
        shop_data[message.chat.id] = {
            'adress': shop_name,  # Сохраняем название магазина
            'cash': None,
            'cashless': None,
            'collection': None,
            'balance': None
        }

        msg = bot.send_message(message.chat.id, f"💰 Введите сумму прихода наличными для магазина '{shop_name}':")
        bot.register_next_step_handler(msg, process_cash)

    except Exception as e:
        bot.send_message(message.chat.id, f"❌ Ошибка: {str(e)}")
        shop_data.pop(message.chat.id, None)

def process_cash(message):
    try:
        chat_id = message.chat.id
        if chat_id not in shop_data:
            raise ValueError("Сессия утеряна. Начните заново командой /report")

        cash = float(message.text.strip())
        shop_data[chat_id]['cash'] = cash
        
        msg = bot.send_message(chat_id, "💳 Теперь введите сумму прихода безналичными:")
        bot.register_next_step_handler(msg, process_cashless)

    except ValueError:
        bot.send_message(chat_id, "❌ Пожалуйста, введите корректную сумму (число)")
    except Exception as e:
        bot.send_message(chat_id, f"❌ Ошибка: {str(e)}")
        shop_data.pop(chat_id, None)

def process_cashless(message):
    try:
        chat_id = message.chat.id
        if chat_id not in shop_data:
            raise ValueError("Сессия утеряна. Начните заново командой /report")

        cashless = float(message.text.strip())
        shop_data[chat_id]['cashless'] = cashless
        
        msg = bot.send_message(chat_id, "🏦 Введите сумму инкассации:")
        bot.register_next_step_handler(msg, process_collection)

    except ValueError:
        bot.send_message(chat_id, "❌ Пожалуйста, введите корректную сумму (число)")
    except Exception as e:
        bot.send_message(chat_id, f"❌ Ошибка: {str(e)}")
        shop_data.pop(chat_id, None)

def process_collection(message):
    try:
        chat_id = message.chat.id
        if chat_id not in shop_data:
            raise ValueError("Сессия утеряна. Начните заново командой /report")

        collection = float(message.text.strip())
        shop_data[chat_id]['collection'] = collection
        
        msg = bot.send_message(chat_id, "📊 Введите остаток в кассе на конец дня:")
        bot.register_next_step_handler(msg, process_balance)

    except ValueError:
        bot.send_message(chat_id, "❌ Пожалуйста, введите корректную сумму (число)")
    except Exception as e:
        bot.send_message(chat_id, f"❌ Ошибка: {str(e)}")
        shop_data.pop(chat_id, None)

def process_balance(message):
    try:
        chat_id = message.chat.id
        if chat_id not in shop_data:
            raise ValueError("Сессия утеряна. Начните заново командой /report")

        balance = float(message.text.strip())
        shop_data[chat_id]['balance'] = balance

        # Проверяем, что все поля заполнены
        required_fields = ['cash', 'cashless', 'collection', 'balance', 'adress']
        for field in required_fields:
            if field not in shop_data[chat_id] or shop_data[chat_id][field] is None:
                raise ValueError(f"❌ Отсутствует значение для поля: {field}")

        # Формируем и отправляем отчет
        data = shop_data[chat_id]
        total_income = data['cash'] + data['cashless']
        total_with_balance = total_income + data['balance']

        report = f"""
📋 Отчет по магазину:
📍 Точка: {data['adress']}
💰 Приход наличными: {data['cash']} руб.
💳 Приход безналичными: {data['cashless']} руб.
🏦 Инкассация: {data['collection']} руб.
📊 Остаток в кассе: {data['balance']} руб.

✅ Итоговый приход: {total_income} руб.
✅ Приход с остатком: {total_with_balance} руб."""

        # Сохраняем отчет
        add_to_df(chat_id)
        
        bot.send_message(chat_id, report)
        bot.send_message(chat_id, "✅ Отчет успешно сохранен!")

    except ValueError as ve:
        bot.send_message(chat_id, str(ve))
    except Exception as e:
        bot.send_message(chat_id, f"❌ Неизвестная ошибка: {str(e)}")
    finally:
        shop_data.pop(chat_id, None)

def add_to_df(chat_id):
    try:
        if chat_id not in shop_data:
            raise ValueError("Данные магазина не найдены")
            
        shop = shop_data[chat_id]
        required_fields = ['adress', 'cash', 'cashless', 'collection', 'balance']
        for field in required_fields:
            if field not in shop or shop[field] is None:
                raise ValueError(f"Отсутствует поле {field}")

        current_datetime = datetime.now()  # Теперь работает правильно
        file_name = f"reports_{current_datetime.year}_{current_datetime.month:02d}.xlsx"
        file_path = os.path.join(REPORTS_DIR, file_name)
        sheet_name = current_datetime.strftime("%Y-%m-%d")

        # Создаем директорию, если ее нет
        os.makedirs(REPORTS_DIR, exist_ok=True)

        # Проверяем существование файла
        if os.path.exists(file_path):
            wb = load_workbook(file_path)
            if sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
            else:
                sheet = wb.create_sheet(title=sheet_name)
                sheet.append(["Дата", "Магазин", "Наличные", "Безналичные", "Инкассация", "Остаток"])
        else:
            wb = Workbook()
            wb.remove(wb.active)  # Удаляем дефолтный лист
            sheet = wb.create_sheet(title=sheet_name)
            sheet.append(["Дата", "Магазин", "Наличные", "Безналичные", "Инкассация", "Остаток"])

        # Добавляем данные
        sheet.append([
            current_datetime.strftime("%Y-%m-%d %H:%M:%S"),
            shop['adress'],
            float(shop['cash']),
            float(shop['cashless']),
            float(shop['collection']),
            float(shop['balance'])
        ])

        # Сохраняем файл
        wb.save(file_path)
        return True

    except PermissionError:
        raise ValueError("Нет прав на запись файла. Закройте файл Excel и попробуйте снова.")
    except Exception as e:
        raise ValueError(f"Ошибка при сохранении в Excel: {str(e)}")
if __name__ == '__main__':
    print('Бот запущен...')
    bot.infinity_polling()